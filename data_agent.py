# Copyright 2022 Carnegie Mellon University Neuromechatronics Lab
#
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.

# Contact: Jonathan Shulgach (jshulgac@andrew.cmu.edu)

import re
import os
import xlrd
import glob
import time
import openpyxl
import scipy.io as sio
import shutil
import datetime
from xlutils.filter import process, XLRDReader, XLWTWriter
import pandas as pd
import numpy as np

from nml_bag.reader import Reader


def save_as(df_file, folder_path, file_name):
    # Helper function to save files in a new directory
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    df_file.to_csv((folder_path + file_name))
    print("Saving complete")


def convert_to_utc(t_val):
    # Helper function to convert ROS2 nanosecond time format to python readable utc time
    # ex: '1661975258802625400' -> 1661975258.802625
    if not isinstance(t_val, str):
        t_val = str(t_val)

    return float(t_val[:10] + '.' + t_val[10:16])


def get_ros2_datatypes(msg):
    # Helper function to get the message sub-datatypes needed 
    if msg == 'rcl_interfaces/msg/Log':
        sub_msg = ['msg', 'file']
    elif msg == 'geometry_msgs/msg/Vector3':
        sub_msg = ['x', 'y', 'z']
    elif msg == 'rcl_interfaces/msg/ParameterEvent':
        sub_msg = ['new_parameters']
    elif msg == 'example_interfaces/msg/String':
        sub_msg = ['data']
    elif msg == 'std_msgs/msg/String':
        sub_msg = ['data']
    elif msg == 'geometry_msgs/msg/Point':
        sub_msg = ['x', 'y', 'z']
    elif msg == 'example_interfaces/msg/Int32':
        sub_msg = ['data']
    elif msg == 'std_msgs/msg/Int32':
        sub_msg = ['data']
    elif msg == 'example_interfaces/msg/Float64':
        sub_msg = ['data']
    elif msg == 'example_interfaces/msg/Bool':
        sub_msg = ['data']
    elif msg == 'std_msgs/msg/ColorRGBA':
        sub_msg = ['r', 'g', 'b', 'a']
    elif msg == 'rosbag2_interfaces/msg/WriteSplitEvent':
        sub_msg = ['opened_file', 'closed_file']
    else:
        print("ROS2 datatype {} not recognized, subclass data unknown or not supported yet".format(msg))
        return None

    return sub_msg


class ArgParser:
    """ An argument parser that searches for keywords in a scanned text and outputs the value associated 
        with the matched key with a filled-in dictionary
    
    Parameters
    ==========
    keys:  (list)
    text_path: (str) path to the directory containing the 'config.txt' file
    delimiter: (str) Delimiter character to split text lines with
    verbose: (bool) enable/disable verbose output
    skip_empty: (bool) enable/diable the option to include keys with empty values
    
    Returns
    =======
    parsed_args: (dict) A dictionary containg keys with each of the parameters discovered in a text file.
    
    """

    def __init__(self, file_path=None, delimiter="=", args=None, skip_line_flag=("#", " ", "\n"), skip_empty=True,
                 verbose=False):
        self.file_path = file_path
        self.delimiter = delimiter
        self.args = args
        self.verbose = verbose
        self.skip_flag = skip_line_flag  # Has to be a single item or tuple
        self.skip_empty = skip_empty
        self.lowercase = True

    def scan_file(self, file_path=None, delimiter=None):
        """ Scans a file for key-value pairs of parameters. Checks the current directory for the file if no directory is specified
        
        Parameters
        ==========
        file_path: (str) (optional) Absolute file path for the file to read, optional parameter if already specified in object
        delimiter: (str) (optional) Delimiter to parse values from keys, optional parameter if already specified in object
        
        Results
        ==========
        parsed_args: (dict) Python dictionary type with the key and value pairs
        """

        if delimiter:
            delim = delimiter
        else:
            delim = self.delimiter

        if self.file_path:
            text_path = self.file_path
        elif file_path:
            text_path = file_path
        else:
            text_path = os.path.join(os.getcwd(), 'config.txt')

        with open(text_path) as f:
            lines = f.readlines()
            # if self.verbose: print(lines)

            parsed_args = {}
            for i, line in enumerate(lines):

                temp = line.split(delim, 1)  # first occurence
                if len(temp) != 2:
                    if self.verbose: print('Parameter missing delimiter on line {}, skipping'.format(i))
                elif temp[1] == '\n' and self.skip_empty:
                    if self.verbose: print('Empty parameter detected, skipping\n')
                elif temp[0].startswith(self.skip_flag):
                    pass  # skip line
                else:
                    val = re.sub('\n', '', temp[1])  # Grab new parameters by checking for new lines
                    # if self.verbose: print(val)
                    if self.lowercase:
                        temp[0] = temp[0].lower()

                    parsed_args[temp[0]] = val
                    if self.verbose:
                        print("Assigning value '", str(val), "' to parameter '", str(temp[0]), "'")

            return parsed_args


def convert_digits_to_timestamp(str_number):
    return '{}:{}:{}'.format(str_number[:2], str_number[2:4], str_number[4:])


def parse_digits_from_string(string):
    """ Internal parser function to extract the date, timestamp, and file block from a file name

    Parameters:
    ----------
    string     : (str) string of filename

    Returns:
    ----------
    date_str: (str) Reformatted string in the date format of 'm/d/yy'
    filetag : (str) timestamp number associated with the file
    block   : (str) File identifier number from the same timestamp recording
    temp    : (str) last match from regular expression

    """
    temp = re.findall('(\d+)', string)
    date_temp = temp[len(temp) - 2]
    time_temp = temp[-1]
    block = 0  # file index at the end
    if len(temp[-1]) <= 2:
        date_temp = temp[len(temp) - 3]
        time_temp = temp[len(temp) - 2]
        block = temp[-1]

    dd = int(date_temp[4:6])
    mm = int(date_temp[2:4])
    yy = int(date_temp[:2])
    date_str = str(mm) + "/" + str(dd) + "/" + str(yy)  # Convenient date format for reading
    datetag = str(date_temp)  # Original 6-digit date format for string matching
    filetag = str(time_temp)  # Original 6-digit time format for string matching

    dt = datetime.date(yy, mm, dd)
    year = '20' + str(dt.year)
    month = dt.month
    day = dt.day

    return datetag, date_str, filetag, block, [year, month, day]


def parse_datestring(string):
    """ Internal parser function to convert a 6-digit date into date components

    Parameters:
    ----------
    string     : (str) 6-digit date string

    Returns:
    ----------
    list    : (list) List with date in [YYYY, MM, DD] format

    """
    temp = re.search('[0-9]{6}', string)
    dt = datetime.date(int(temp[0][:2]), int(temp[0][2:4]), int(temp[0][4:6]))
    year = '20' + str(dt.year)
    month = dt.month
    day = dt.day
    return [year, month, day]


def parse_str_date_info(string, year_format=None):
    """ Internal parser function to extract the date from a file name

    Parameters:
    ----------
    string     : (str) string of filename

    Returns:
    ----------
    info_str: (str) Reformatted string in the generated_data folder date format
    temp    : (str) last match from regular expression

    """
    temp = re.search('[0-9]{6}', string)
    dd = int(temp[0][4:6])
    mm = int(temp[0][2:4])
    yy = int(temp[0][:2])
    info_str = str(mm) + "/" + str(dd) + "/" + str(yy)

    return info_str, temp.group()


class DataAgent:
    """ A data sorting agent that searches a directory for matching file types and extracts file data with keyword inputs.

    Parameters
    ----------
    *args :      (list) system arguments passed through to the constructor
    subject:     (str) Name of the subject to identify data with
    search_path: (str) Directory for raw data files
    save_path:   (str) Directory for the generated data files
    param_path:  (str) directory for the task's .yaml config files for metadata collection
    file_type:   (str) file type to search for in all folders from root directory
    verbose:     (bool) enable/disable verbose output
    *kargs :     (dict) Keyword arguments passed to superclass constructor_.
    """

    def __init__(self,
                 subject=None,
                 search_path=None,
                 save_path=None,
                 param_path=None,
                 file_type=".db3",
                 notebook_path=r'',
                 verbose=False,
                 **kwargs):
        self.verbose = verbose
        self.file_type = file_type
        self.subject = subject
        self.search_path = search_path
        self.save_path = save_path
        self.param_path = param_path
        self.notebook_path = notebook_path
        self.file_list = []
        self.date_list = []
        self.date = []
        self._data = []
        self.notebook_headers = None
        self.notebook_file = None
        self.notebook_copy = None
        self.notebook_style_list = None

        # Check that the paths exist
        if self.search_path:
            self.check_path(self.search_path)
        if self.save_path:
            self.check_path(self.save_path)
        if self.param_path:
            self.check_path(self.param_path)

    def add_param_path(self, param_path=None):
        """ Adds a search directory for parameters saved in config files 
        """
        if os.path.isdir(param_path):
            self.param_path = param_path

    def build_path(self, date, year_first=True):
        """ Helper function that creates folder subdirectories based on the date passed into the 
            proper subject and date-specified folders.
        
        Parameters:
        ----------
        date       : (str)  String containing a date entry in the format of MM/DD/YY
        year_first : (bool) Boolean that switches the position of the year in the date entry 
                              to appear at the beginning instead of the end
                                
        Return
        ============
        save_folder_path  : (str) The absolute have to the saved data directory including the 
                             subject and modified date subdirectories
        new_date          : (str) The passed date with the new string format 
            
            
        Example:
        ============
        >> agent = DataAgent('/home/user/documents')
        >> save_path, new_date = agent.build_path('7/27/23', 'TestUser')
        >> print(save_path)
        >> '/home/user/documents/TestUser/TestUser_2023_07_27'
             
        """
        date_parts = date.split('/')
        if year_first:
            new_date = '20' + date_parts[2] + '_' + date_parts[0].zfill(2) + '_' + date_parts[1].zfill(2)
            # new_date =  '_'.join(['20'+date_parts[2]] + date_parts[:2])
        else:
            new_date = date_parts[0].zfill(2) + '_' + date_parts[1].zfill(2) + '_' + '20' + date_parts[2]
            # new_date = '_'.join(date_parts[:2]+['20'+date_parts[2]])
        if self.save_path:
            parent_folder = self.save_path
        else:
            parent_folder = ''
        save_folder_path = parent_folder + '\\' + self.subject + '\\' + self.subject + '_' + new_date + '\\'
        return save_folder_path, new_date

    def check_path(self, data_path=None):
        """Checks the folder directory as a valid path and looks for bag files in subdirectories
        
        Parameters:
        ----------
        data_path : (str) Absolute path directory to folder
        
        Return:
        ----------
        data_path : (str, None) Returns the same data path provided if valid path, otherwise None
        
        """
        if data_path is None:
            data_path = input("Please enter the parent folder directory: ")

        if os.path.isdir(data_path):
            if self.verbose: print('Valid root directory')
            return data_path
        else:
            print("Warning, path {} doesn't exist. Returning 'None':\n".format(data_path))
            return None

    def check_notebook_entry(self, table, date_str, headers, overwrite=False):
        """ Check which headers are empty for new date entry, if not then skip
        
        Parameters:
        ----------
        table     :
        date_str  :
        headers   :
        overwrite :
        
        Return:
        ----------
        skip      : (list) A list of boolean values indexed by the argument header
        
        """
        if headers is None:
            print("Warning, no header field names specified. Returning 'None'")
            return None
        else:
            skip = [False for i in range(len(headers))]
            for h, header in enumerate(headers):
                if isinstance(table, pd.DataFrame):
                    if header not in table.columns:
                        if self.verbose: print("Warning - {} not found in data frame columns".format(header))
                    else:
                        date_idx = table.index[table[table.columns[0]] == date_str].tolist()[0]
                        if not np.isnan(table.loc[date_idx, header]):
                            if self.verbose: print("Warning - {} already contains data. Skipping")
                            if not overwrite:
                                skip[h] = True
                elif isinstance(table, xlrd.book.Book):
                    pass
                    # if header not in

                # if isinstance(table, Workbook):
                #    # Only supporting dataframes for now
                #    pass

            return skip

    def find_notebook_columns(self, headers, notebook_sheet=None, is_xlsx=True):
        """ Function that searches a dataframe or Workbook object for the indices of the specified headers and
        returns them. If a list of headers is passed, the function will return a list of indices for each header in
        the list. If the header does not exist, the element will contain a -1

        Parameters:
        ----------
        headers : (list) List of strings containing the header names to search for
        notebook_sheet : (str) Name of the sheet in the workbook to search for the headers
        is_xlsx : (bool) Option to specify if the workbook is in .xlsx format

        Returns:
        ----------
        indices : (list) List of integers containing the indices of the headers in the dataframe or workbook object

        """
        if isinstance(self.notebook_file, pd.DataFrame):
            return [self.notebook_file.columns.get_loc(header) for header in headers]

        elif isinstance(self.notebook_file, openpyxl.workbook.workbook.Workbook):
            if notebook_sheet is None and self.notebook_sheet is None:
                print(" - Warning - No sheet specified. Returning 'None'")
                return None
            elif notebook_sheet is None:
                notebook_sheet = self.notebook_sheet

            ws_hdrs = [cell.value for cell in self.notebook_file[notebook_sheet][1]]

            # Remove spaces at the beginning and end of the header names in ws_hdrs
            ws_hdrs = [header.strip() for header in ws_hdrs if header is not None]

            # If header is in ws_headers save the index + 1 (to match 1-indexed excel format)
            # If not, remember that, so headers will return only headers found in ws_hdrs and indices is the same length
            indices = []
            final_headers = []
            for i, header in enumerate(headers):
                if header in ws_hdrs:
                    indices.append(ws_hdrs.index(header) + 1)
                    final_headers.append(header)
                else:
                    print(" - Warning - {} not found in workbook".format(header))

            return dict(zip(final_headers, indices))



        elif isinstance(self.notebook_file, xlrd.book.Book):
            if notebook_sheet is None and self.notebook_sheet is None:
                print("Warning - No sheet specified. Returning 'None'")
                return None
            elif notebook_sheet is None:
                notebook_sheet = self.notebook_sheet
            return [self.notebook_file.sheet_by_name(notebook_sheet).row_values(0).index(header) for header in headers]

        else:
            print("Warning - Data type not recognized. Returning 'None'")
            return None

    def find_notebook_row(self, row_name, notebook_sheet=None):
        """ Function that searches a dataframe or Workbook object for the row index of a row_name (usually a date in
        the format of DD/MM/YY) and returns it. The first column contains dates in the same format as the row_name

        Parameters
        ----------
        row_name : (str) String containing the row name to search for
        notebook_sheet : (str) Name of the sheet in the workbook to search for the row

        Returns
        ----------
        index : (int) Integer containing the index of the row in the dataframe or workbook object
        """
        if isinstance(self.notebook_file, pd.DataFrame):
            return self.notebook_file.index[self.notebook_file[self.notebook_file.columns[0]] == row_name].tolist()[0]
        elif isinstance(self.notebook_file, xlrd.book.Book):
            if notebook_sheet is None and self.notebook_sheet is None:
                print("Warning - No sheet specified. Returning 'None'")
                return None
            elif notebook_sheet is None:
                notebook_sheet = self.notebook_sheet

            split_date = row_name.split('/')
            if split_date[0][0] == '0':  # Remove leading zero from day or month
                split_date[0] = split_date[0][1]
            if split_date[1][0] == '0':
                split_date[1] = split_date[1][1]

            return self.notebook_file.sheet_by_name(notebook_sheet).col_values(0).index(row_name)
        elif isinstance(self.notebook_file, openpyxl.workbook.workbook.Workbook):
            if notebook_sheet is None and self.notebook_sheet is None:
                print("Warning - No sheet specified. Returning 'None'")
                return None
            elif notebook_sheet is None:
                notebook_sheet = self.notebook_sheet

            for row in self.notebook_file[notebook_sheet].iter_rows(min_col=1, max_col=1):
                row_val = row[0].value
                if isinstance(row_val, datetime.datetime):
                    row_val = row_val.strftime('%m/%d/%y')

                if row_val == row_name:
                    return row[0].row

        else:
            print("Warning - Data type not recognized. Returning 'None'")
            return None

    def load_notebook(self, file_path, use_df=False, keep_formatting_info=True, return_notebook=False, is_xlsx=True):
        """ Load a workbook file and return the data as a dataframe or workbook object

        Parameters
        ----------
        file_path: (str) Absolute path to the workbook file
        use_df: (bool) Option to return the data as a dataframe or workbook object
        keep_formatting_info: (bool) Option to return the formatting information from the workbook
        return_notebook: (bool) Option to return the workbook object

        Returns
        -------
        notebook: (DataFrame, Book) pandas dataframe or book object containing the data from the loaded file
        """
        if not os.path.isfile(file_path):
            print("Error - path not valid: {}".format(file_path))
            return
        else:
            if self.verbose: print("Found notebook '{}'".format(file_path))
        if use_df:
            self.notebook_file = pd.read_excel(file_path, None)
        else:
            if is_xlsx:
                self.notebook_file = openpyxl.load_workbook(file_path)
            else:
                self.notebook_file = xlrd.open_workbook(file_path, formatting_info=keep_formatting_info)

        if False:
            # Make a copy
            if isinstance(self.notebook_file, xlrd.book.Book):
                if self.notebook_copy is None:
                    #  Copy the file
                    w = XLWTWriter()
                    process(XLRDReader(self.notebook_file, 'unknown.xls'), w)
                    self.notebook_copy = w.output[0][1]
                    self.notebook_style_list = w.style_list
            elif isinstance(self.notebook_file, openpyxl.workbook.workbook.Workbook):
                # Create a new workbook
                self.notebook_copy = openpyxl.Workbook()

                # Iterate through each sheet in the original workbook
                for original_sheet in self.notebook_file.sheetnames:
                    original_ws = self.notebook_file[original_sheet]
                    new_ws = self.notebook_copy.create_sheet(title=original_sheet)

                    # Iterate through each cell in the original sheet
                    for row in original_ws.iter_rows(values_only=True):
                        new_ws.append(row)

                    # Copy styles from original to new workbook
                    for row in range(1, original_ws.max_row + 1):
                        for column in range(1, original_ws.max_column + 1):
                            source_cell = original_ws.cell(row=row, column=column)
                            target_cell = new_ws.cell(row=row, column=column)
                            target_cell._style = source_cell._style

        if return_notebook:
            return self.notebook_file

    def get_cursor_pos(self, date, topic='/environment/cursor/position', save=False, file_type='txt', overwrite=False):
        """ Helper function that loads the cursor position data from a loaded bag file and stores it 
            to a local file (.txt by default)
        
        Parameters
        ----------
        date       : (str) String containing a date entry in the format of MM/DD/YY
        topic      : (str) ROS2 topic containing the information
        file_type  : (str) Date type of the log file
        save       : (bool) Option to save the data after loading
        overwrite  : (bool) Option to overwrite existing file
        
        """

        [folder_path, new_date] = self.build_path(date)
        file_name = new_date + '_CURSOR_POS.' + file_type

        if os.path.isfile(os.path.join(folder_path, file_name)):
            print(
                "File '{}' in directory '{}' already exists. Overwriting not supported. Please delete the file and "
                "resave".format(
                    file_name, folder_path))
            return

        print('Grabbing cursor position data...\n')
        df_file = self.get_topic_data(topic)  # For just cursor position

        if save:
            self.save_file(df_file, folder_path, file_name)

        print("Done")

    def get_states(self, date, topic='/machine/state', save=False, file_type='txt', overwrite=False):
        """ Helper function that loads the task states from a loaded bag file and stores it to a 
            local file (.txt by default)
        
        Parameters
        ----------
        date       : (str) String containing a date entry in the format of MM/DD/YY
        topic      : (str) ROS2 topic containing the information
        file_type  : (str) Date type of the log file
        save       : (bool) Option to save the data after loading
        overwrite  : (bool) Option to overwrite existing file
        
        """

        [folder_path, new_date] = self.build_path(date)
        file_name = new_date + '_STATES.' + file_type
        # print(os.path.join(folder_path, file_name))
        if os.path.isfile(os.path.join(folder_path, file_name)):
            print(
                "File '{}' in directory '{}' already exists. Overwriting not supported. Please delete the file and "
                "resave".format(
                    file_name, folder_path))
            return

        print('Grabbing states...\n')
        df_file = self.get_topic_data(topic)  # For just task states

        if save:
            self.save_file(df_file, folder_path, file_name)

        print("Done")

    def get_emg(self, date, topic='/pico_ros/emg/ch1', save=False, file_type='txt', overwrite=False):
        """ Helper function that loads the electromyographic recordings of a specific channel from a 
            loaded bag file and stores it to a local file (.txt by default)
        
        Parameters:
        ----------
        date       : (str) String containing a date entry in the format of MM/DD/YY
        topic      : (str) ROS2 topic containing the information
        file_type  : (str) Date type of the log file
        save       : (bool) Option to save the data after loading
        overwrite  : (bool) Option to overwrite existing file
        
        """
        [folder_path, new_date] = self.build_path(date)
        file_name = new_date + '_EMG.' + file_type

        if os.path.isfile(os.path.join(folder_path, file_name)):
            print(
                "File '{}' in directory '{}' already exists. Overwriting not supported. Please delete the file and "
                "resave".format(
                    file_name, folder_path))
            return

        print('Grabbing EMG data...\n')
        df_file = self.get_topic_data(topic)  # For just task states

        if save:
            self.save_file(df_file, folder_path, file_name)

        print("Done")

    def get_targets(self, date, topic='/environment/target/position', save=False, file_type='txt', overwrite=False):
        """ Helper function that loads the target position data from a loaded bag file and stores it 
            to a local file (.txt by default)
        
        Parameters:
        ----------
        date       : (str) String containing a date entry in the format of MM/DD/YY
        topic      : (str) ROS2 topic containing the information
        file_type  : (str) Date type of the log file
        save       : (bool) Option to save the data after loading
        overwrite  : (bool) Option to overwrite existing file
        """

        [folder_path, new_date] = self.build_path(date)
        file_name = new_date + '_TARGETS.' + file_type
        if os.path.isfile(os.path.join(folder_path, file_name)):
            print(
                "File '{}' in directory '{}' already exists. Overwriting not supported. Please delete the file and "
                "resave".format(
                    file_name, folder_path))
            return

        print('Grabbing targets...\n')
        df_file = self.get_topic_data(topic)  # For just task states

        if save:
            self.save_file(df_file, folder_path, file_name)

        print("Done")

    def get_forces(self, date, file_type='txt', save=False, overwrite=False):
        """ Function that loads the force transformation data for the robot and the cursor from a 
            loaded bag file and stores it to a local file (.txt by default)
        
        Parameters:
        ----------
        date       : (str) String containing a date entry in the format of MM/DD/YY
        file_type  : (str) Date type of the log file
        save       : (bool) Option to save the data after loading
        overwrite  : (bool) Option to overwrite existing file
            
        """
        [folder_path, new_date] = self.build_path(date)
        file_name_1 = new_date + '_FORCE_ROBOT_FEEDBACK.' + file_type
        file_name_2 = new_date + '_FORCE_ROBOT_COMMAND.' + file_type
        file_name_3 = new_date + '_FORCE_CURSOR.' + file_type
        if os.path.isfile(os.path.join(folder_path, file_name_1)):
            print(
                "File '{}' in directory '{}' already exists. Overwriting not supported. Please delete the file and "
                "resave".format(
                    file_name_1, folder_path))
            return

        print('Grabbing force data...\n')
        f_robot_df_file = self.get_topic_data('/robot/feedback/force')  # Force feedback from robot
        f_cmd_df_file = self.get_topic_data('/robot/command/force')  # Force commands sent to robot
        f_cursor_df_file = self.get_topic_data('/cursor/force')  # Cursor force feedback

        if save:
            print('Saving force data to:\n ' + (
                    folder_path + new_date + '_FORCE_* custom topic extension:\n  "ROBOT_FEEDBACK"\n  '
                                             '"ROBOT_COMMAND"\n  "CURSOR\n")'))
            save_as(f_robot_df_file, folder_path, file_name_1)
            save_as(f_cmd_df_file, folder_path, file_name_2)
            save_as(f_cursor_df_file, folder_path, file_name_3)
            print("Done")

    def get_metrics(self, date, data=None, state_topic='/machine/state', display_metrics=False, save=False,
                    file_type='txt', overwrite=False, return_metrics=False):
        """ Saves the performance metrics of the loaded database in the directory specified as 
            a .txt file by default unless also specified
        
        Parameters
        ----------
        date            : (str) String containing a date entry in the format of MM/DD/YY
        state_topic           : (str) ROS2 topic containing the information
        data            : (DataFrame) Optional Dataframe containing the loaded data
        file_type       : (str) Date type of the log file
        save            : (bool) Option to save the data after loading
        overwrite       : (bool) Option to overwrite existing file
        display_metrics : (bool) Option to display the metrics on the screen
        return_metrics  : (bool) Option to return the metrics as a dictionary

        """
        if data is not None and isinstance(data, pd.DataFrame):
            self._data = data

        [folder_path, new_date] = self.build_path(date)
        file_name = new_date + '_PERFORMANCE_METRICS.' + file_type

        if os.path.isfile(os.path.join(folder_path, file_name)) and not overwrite:
            print("File '{}' in directory '{}' already exists and overwriting set to False. Please set overwrite to "
                  "True or delete the file".format(file_name, folder_path))
            return

        # Get TOTAL TRIALS, CORRECT TRIALS, START TIME, TIME WORKED
        print('Grabbing performance metrics...')
        trial_perf = self.get_trial_performance(state_topic)
        avg_trial_t = self.get_mean_trial_time(state_topic)
        if trial_perf['TOTAL TRIALS'] is None:
            print(" | Failed to get performance metrics")
            return

        # Get START TIME and TIME WORKED
        time_ns = self._data['time_ns']
        # START_TIME = time.strftime('%Y-%m-%d %H:%M:%S.%f', time.localtime(time_ns.iloc[0] / 1e9))
        t = datetime.datetime.utcfromtimestamp(time_ns.iloc[0] / 1e9)
        START_TIME = t.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

        # Get the number of minutes that passed between the first and last time_ns
        TIME_WORKED = (time_ns.iloc[-1] - time_ns.iloc[0]) / 1e9 / 60

        # Decided to add parameters to metrics file, don't need to make new file to parse yet
        if not self.param_path:
            print('Warning - Parameter file path not set. USe the add_param_path() function with the directory holding '
                  'your parameter files.')
        else:
            n_targets = self.get_param_from_file(self.param_path, 'n_targets')
            target_r = self.get_param_from_file(self.param_path, ['target', 'radius'])
            cursor_r = self.get_param_from_file(self.param_path, ['cursor', 'radius'])
            # enforce_orient = self.get_param_from_file(self.param_path, 'enforce_orientation')

        # Display the trial performance metrics on screen
        if display_metrics:
            print("=======================================")
            print("[{}] Performance metrics:\n".format(new_date))
            print("Total number of trials: {}".format(trial_perf['TOTAL TRIALS']))
            print("Correct trials:         {}".format(trial_perf['CORRECT TRIALS']))
            print("Percentage correct:     {:.1f}".format(
                100 * trial_perf['CORRECT TRIALS'] / trial_perf['TOTAL TRIALS']))
            print("Average trial time:     {:.2f} s".format(avg_trial_t))
            print("Start time:             {}".format(START_TIME))
            print("Time worked:            {:.2f} minutes".format(TIME_WORKED))
            print("=======================================")

        if save:
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)

            print('Saving metrics data to "' + (folder_path + file_name) + '"')
            with open((folder_path + file_name), "w") as f:

                # ===== Add any metadata information you want to save here ========
                msg = new_date + ",N:" + str(trial_perf['TOTAL TRIALS']) + ",Success:" + str(
                    trial_perf['CORRECT TRIALS']) + ",Failure:" + str(
                    trial_perf['TOTAL TRIALS'] - trial_perf['CORRECT TRIALS']) + ",Success_Rate:" + str(
                    100 * trial_perf['CORRECT TRIALS'] / trial_perf['TOTAL TRIALS']) + "\n"
                f.write(msg)
                f.write('START_TIME:' + START_TIME + "\n")
                f.write('TIME_WORKED:' + str(TIME_WORKED) + "\n")

                if avg_trial_t is not None:
                    f.write('MEAN_TRIAL_T:' + str(avg_trial_t) + "\n")

                if self.param_path:
                    if n_targets is not None:
                        f.write('N_TARGETS:' + n_targets + "\n")
                    if target_r is not None:
                        f.write('TARGET_RADIUS:' + target_r + "\n")
                    if cursor_r is not None:
                        f.write('CURSOR_RADIUS:' + cursor_r + "\n")
                    # if enforce_orient is not None:
                    #    f.write('ENFORCE_ORIENTATION:' + enforce_orient + "\n"

                f.close()

            if return_metrics:
                # 'LIQUID EARNED', 'FREE WATER', 'TOTAL TRIALS', 'CORRECT TRIALS', 'START TIME', 'TIME WORKED'
                # Note: these have to match the header names in the notebook file
                data = {'TOTAL TRIALS': trial_perf['TOTAL TRIALS'],
                        'CORRECT TRIALS': trial_perf['CORRECT TRIALS'],
                        'PERCENT CORRECT': "{:.1f}".format(
                            100 * trial_perf['CORRECT TRIALS'] / trial_perf['TOTAL TRIALS']),
                        'AVERAGE TRIAL TIME': "{:.1f}".format(trial_perf['AVERAGE TRIAL TIME']),
                        'START TIME': START_TIME,
                        'TIME WORKED (mins)': "{:.1f}".format(TIME_WORKED),
                        'PRIMARY TARGET MOVE ERROR': trial_perf['PRIMARY TARGET MOVE ERROR'],
                        'PRIMARY TARGET OVERSHOOT': trial_perf['PRIMARY TARGET OVERSHOOT'],
                        'PRIMARY TARGET OVERSHOOT ERROR': trial_perf['PRIMARY TARGET OVERSHOOT ERROR'],
                        'SECONDARY TARGET INSTRUCTION ERROR': trial_perf['SECONDARY TARGET INSTRUCTION ERROR'],
                        'SECONDARY TARGET MOVE ERROR': trial_perf['SECONDARY TARGET MOVE ERROR'],
                        'SECONDARY TARGET OVERSHOOT': trial_perf['SECONDARY TARGET OVERSHOOT'],
                        'SECONDARY TARGET OVERSHOOT ERROR': trial_perf['SECONDARY TARGET OVERSHOOT ERROR'],
                        'PRIMARY TARGET RETURN INSTRUCTION ERROR': trial_perf['PRIMARY TARGET RETURN INSTRUCTION ERROR'],
                        'PRIMARY TARGET RETURN MOVE ERROR': trial_perf['PRIMARY TARGET RETURN MOVE ERROR'],
                        'PRIMARY TARGET RETURN OVERSHOOT': trial_perf['PRIMARY TARGET RETURN OVERSHOOT'],
                        'PRIMARY TARGET RETURN OVERSHOOT ERROR': trial_perf['PRIMARY TARGET RETURN OVERSHOOT ERROR'],
                        'SUCCESS WITH OVERSHOOT': trial_perf['SUCCESS WITH OVERSHOOT'],
                        }
                if self.param_path:
                    data['N_TARGETS'] = n_targets
                    data['TARGET_RADIUS'] = target_r
                    data['CURSOR_RADIUS'] = cursor_r
                    # data['ENFORCE_ORIENTATION'] = enforce_orient
                return data

    def get_param_from_file(self, param_path=None, param_name=None, file_type='yaml'):
        """ Checks directory for config files in .yaml format, searches for the specified 
            parameter and returns an associated value
        
        Parameters:
        ----------
        param_path : (str) directory for config files
        param_name : (str/list) Single or list of strings of parameter name (and subfields) to search
        file_type  : (str) The file format to look specifically for (default: yaml)
            
        Return:
        ---------
        param_val : (str) value associated with the parameter

        """

        if param_name is None:
            print("Please pass parameter name to search for and try again")
            return

        if type(param_name) is not list:
            param_name = [param_name]  # Enforce being in a list for building regex search expression

        param_val = None
        file_list = glob.glob(self.param_path + "*/*." + file_type)
        if self.verbose:
            print("Found files: \n")
            [print("{}".format(ff)) for ff in file_list]

        # Check each file for the parameter until found
        param_found = False
        while not param_found:
            for file in file_list:
                with open(file, encoding='utf-8') as f:
                    text_output = f.read()
                    temp = [p + r':\s+' for p in param_name]
                    str_exp = "(" + ''.join(temp) + ")([\S]*)"
                    match = re.findall(str_exp, text_output)
                if match:
                    if self.verbose: print(
                        "Found parameter sequence '{}' in file '{}'. Getting value...".format(param_name, file))
                    param_val = match[0][1]
                    param_found = True
                    return param_val
                else:
                    if self.verbose: print("Could not find parameter '{}' in file '{}'".format(param_name, file))

        if param_val is None:
            print("Could not find parameter '{}' in directory {}'".format(param_name, param_path))

    def get_topic_data(self, topic, options=None):
        """ Function that returns a datafram object containing the topic's data if present
        
        Parameters:
        ----------
        bag_topic       : (str) topic to access
        options         : (list) list containing keywords for data types ('cursor' | 'state') TO-DO
   
        Return:
        -----------
        parsed_topic_df : (Dataframe) pandas dataframe table of parsed topic data

        Notes
        --------
        The 'options' parameter can be expanded for more unique data types as DataAgent evolves
        
        """

        if not isinstance(topic, str):
            print("Warning: topic input must be string")
            return None

        # Check that the first character in the topic starts with a '/'
        if topic[0] != '/':
            topic = '/' + topic

        if self.verbose:
            print("Searching for '{}' topic data...00%".format(topic), end="")
        # print("\b\b\b{:02d}%".format(val), end="")
        # for i, reader in enumerate(self._data):
        # print("\b\b\b{:02d}%".format(int(100 * (i + 1) / (len(self._data) + 1))))

        # Return each row in the dataframe where the contents of the 'topic' column match the topic input
        parsed_topic_df = None
        if 'topic' not in self._data.columns:
            print("Warning, topic '{}' not found in bag file, Stopping search.".format(topic))
        else:
            parsed_topic_df = self._data.loc[self._data['topic'] == topic]

        if self.verbose:
            print("\b\b\bDone")

        return parsed_topic_df

    def get_trial_performance(self, state_topic='/machine/state'):
        """ Helper function that gets the task performance statistics. Reads the state topic to 
            evaluate performance (default: /a\machine/state)
        
        Parameters:
        -----------
        state_topic     : (str) The ROS2 topic to read task states

        Returns
        ---------
        data  : (dict) Dictionary containing the performance metrics. The full list of metrics is:
                - CORRECT TRIALS
                - INCORRECT TRIALS
                - TOTAL TRIALS
                - AVERAGE TRIAL TIME
        
        """
        state_df = self.get_topic_data(state_topic)

        # Remove duplicate states
        prev_state = None
        duplicates_to_ignore = ['intertrial', 'move_a']
        for idx, row in state_df.iterrows():
            current_state = row['data']
            if current_state == prev_state and current_state in duplicates_to_ignore:
                state_df.at[idx, 'data'] = ''
            prev_state = current_state

        # Removing empty strings results in duplicate states removed
        state_df = state_df[state_df['data'] != '']

        # Reset indices
        state_df.reset_index(drop=True, inplace=True)

        # Getting some state indices
        success_idx_list = state_df.index[state_df['data'] == 'success']
        failure_idx_list = state_df.index[state_df['data'] == 'failure']
        move_a_idx_list = state_df.index[state_df['data'] == 'move_a']
        move_b_idx_list = state_df.index[state_df['data'] == 'move_b']
        move_c_idx_list = state_df.index[state_df['data'] == 'move_c']
        hold_a_idx_list = state_df.index[state_df['data'] == 'hold_a']
        hold_b_idx_list = state_df.index[state_df['data'] == 'hold_b']
        hold_c_idx_list = state_df.index[state_df['data'] == 'hold_c']
        intertrial_idx_list = state_df.index[state_df['data'] == 'intertrial']
        overshoot_a_idx_list = state_df.index[state_df['data'] == 'overshoot_a']
        overshoot_b_idx_list = state_df.index[state_df['data'] == 'overshoot_b']
        overshoot_c_idx_list = state_df.index[state_df['data'] == 'overshoot_c']
        delay_a_idx_list = state_df.index[state_df['data'] == 'delay_a']
        delay_b_idx_list = state_df.index[state_df['data'] == 'delay_b']
        delay_c_idx_list = state_df.index[state_df['data'] == 'delay_c']

        # Begin collecting metrics
        data = {}
        if state_df is not None:
            # CORRECT TRIALS is the total number of success states in the data
            data['CORRECT TRIALS'] = len(state_df[state_df['data'] == 'success'])

            # INCORRECT TRIALS is the total number of failure states in the data
            data['INCORRECT TRIALS'] = len(state_df[state_df['data'] == 'failure'])

            # TOTAL TRIALS is the sum of correct and incorrect trials
            data['TOTAL TRIALS'] = data['CORRECT TRIALS'] + data['INCORRECT TRIALS']

            # SUCCESS WITH OVERSHOOT is the number of trial periods between the move_a and success states where an
            # overshoot state is present. We can find this while getting the average trial time
            data['SUCCESS WITH OVERSHOOT'] = 0
            # AVERAGE TRIAL TIME is the average time difference between the move_a state and following success state
            time_diff = []
            for success_idx in success_idx_list:
                prev_move_a_idx = [x for x in move_a_idx_list if x < success_idx][-1]
                if prev_move_a_idx is not None:
                    t_1 = convert_to_utc(state_df.loc[prev_move_a_idx, 'time_ns'])
                    t_0 = convert_to_utc(state_df.loc[success_idx, 'time_ns'])
                    time_diff.append(t_0 - t_1)

                    # While we have this period, check if the "overshoot_a", "overshoot_b", or "overshoot_c" states
                    # are present between these indices
                    if len(overshoot_a_idx_list) > 0:
                        for i in range(len(overshoot_a_idx_list)):
                            if prev_move_a_idx < overshoot_a_idx_list[i] < success_idx:
                                data['SUCCESS WITH OVERSHOOT'] += 1
                    if len(overshoot_b_idx_list) > 0:
                        for i in range(len(overshoot_b_idx_list)):
                            if prev_move_a_idx < overshoot_b_idx_list[i] < success_idx:
                                data['SUCCESS WITH OVERSHOOT'] += 1
                    if len(overshoot_c_idx_list) > 0:
                        for i in range(len(overshoot_c_idx_list)):
                            if prev_move_a_idx < overshoot_c_idx_list[i] < success_idx:
                                data['SUCCESS WITH OVERSHOOT'] += 1

            time_diff = [x for x in time_diff if x <= 10]  # Remove values greater than 10 seconds
            if len(time_diff) > 1:
                data['AVERAGE TRIAL TIME'] = np.mean(time_diff)

            # PRIMARY TARGET MOVE ERROR is the number of times the next state after move_a is failure
            data['PRIMARY TARGET MOVE ERROR'] = 0
            for i in range(1, len(failure_idx_list)):
                if state_df['data'].iloc[failure_idx_list[i] - 1] == "move_a":
                    data['PRIMARY TARGET MOVE ERROR'] += 1

            # PRIMARY TARGET OVERSHOOT ERROR is the number of times the next state after hold_a is overshoot_a
            data['PRIMARY TARGET OVERSHOOT'] = 0
            for i in range(1, len(overshoot_a_idx_list)):
                if state_df['data'].iloc[overshoot_a_idx_list[i] - 1] == "hold_a":
                    data['PRIMARY TARGET OVERSHOOT'] += 1

            # SECONDARY TARGET INSTRUCTION ERROR is the number of times the next state after delay_a is failure
            data['SECONDARY TARGET INSTRUCTION ERROR'] = 0
            for i in range(1, len(failure_idx_list)):
                if state_df['data'].iloc[failure_idx_list[i] - 1] == "delay_a":
                    data['SECONDARY TARGET INSTRUCTION ERROR'] += 1

            # PRIMARY TARGET OVERSHOOT ERROR is the number of times the next state after overshoot_a is error
            data['PRIMARY TARGET OVERSHOOT ERROR'] = 0
            for i in range(1, len(failure_idx_list)):
                if state_df['data'].iloc[failure_idx_list[i] - 1] == "overshoot_a":
                    data['PRIMARY TARGET OVERSHOOT ERROR'] += 1

            # SECONDARY TARGET MOVE ERROR is the number of times the next state after move_b is failure
            data['SECONDARY TARGET MOVE ERROR'] = 0
            for i in range(1, len(failure_idx_list)):
                if state_df['data'].iloc[failure_idx_list[i] - 1] == "move_b":
                    data['SECONDARY TARGET MOVE ERROR'] += 1

            # SECONDARY TARGET OVERSHOOT is the number of times the next state after hold_b is overshoot_b
            data['SECONDARY TARGET OVERSHOOT'] = 0
            for i in range(1, len(overshoot_b_idx_list)):
                if state_df['data'].iloc[overshoot_b_idx_list[i] - 1] == "hold_b":
                    data['SECONDARY TARGET OVERSHOOT'] += 1

            # PRIMARY TARGET RETURN INSTRUCTION ERROR is the number of times the next state after delay_b is failure
            data['PRIMARY TARGET RETURN INSTRUCTION ERROR'] = 0
            for i in range(1, len(failure_idx_list)):
                if state_df['data'].iloc[failure_idx_list[i] - 1] == "delay_b":
                    data['PRIMARY TARGET RETURN INSTRUCTION ERROR'] += 1

            # SECONDARY TARGET OVERSHOOT ERROR is the number of times the next state after overshoot_b is error
            data['SECONDARY TARGET OVERSHOOT ERROR'] = 0
            for i in range(1, len(failure_idx_list)):
                if state_df['data'].iloc[failure_idx_list[i] - 1] == "overshoot_b":
                    data['SECONDARY TARGET OVERSHOOT ERROR'] += 1

            # PRIMARY TARGET RETURN MOVE ERROR is the number of times the next state after move_c is failure
            data['PRIMARY TARGET RETURN MOVE ERROR'] = 0
            for i in range(1, len(failure_idx_list)):
                if state_df['data'].iloc[failure_idx_list[i] - 1] == "move_c":
                    data['PRIMARY TARGET RETURN MOVE ERROR'] += 1

            # PRIMARY TARGET RETURN OVERSHOOT is the number of times the next state after hold_c is overshoot_c
            data['PRIMARY TARGET RETURN OVERSHOOT'] = 0
            for i in range(1, len(overshoot_c_idx_list)):
                if state_df['data'].iloc[overshoot_c_idx_list[i] - 1] == "hold_c":
                    data['PRIMARY TARGET RETURN OVERSHOOT'] += 1

            # PRIMARY TARGET RETURN OVERSHOOT ERROR is the number of times the next state after overshoot_c is error
            data['PRIMARY TARGET RETURN OVERSHOOT ERROR'] = 0
            for i in range(1, len(failure_idx_list)):
                if state_df['data'].iloc[failure_idx_list[i] - 1] == "overshoot_c":
                    data['PRIMARY TARGET RETURN OVERSHOOT ERROR'] += 1

        return data

    def get_mean_trial_time(self, topic='task/state', start_state='move_a'):
        """ Function that finds the average duration of trials from the task state changes
        
        Parameters:
        -----------
        topic              : (str) The ROS2 topic to read task states
        start_state        : (str) Starting state name to reference trial times with

        Returns
        ---------
        mean_total_trial_t : (float)
        
        Note:
          Not sure if this will be the more accurate way to do it, but another method below:
           - In loop for length of data
           - Grab indices of success states from last to first
           - find index of first occurence of hold_a state. 
           - Get time difference between states
           - Find next success state continue
           - subtract SUCCESS state time with MOVE_A state time
          
        """

        # For now we find the time between all hold_a states
        df_states = self.get_topic_data(topic)  # For just task states
        if df_states is not None:
            start_idx = df_states.index[df_states['data'] == start_state].tolist()
            start_t = [convert_to_utc(i) for i in df_states['time_ns'][start_idx]]
            mean_total_trial_t = np.mean(np.diff(start_t))
        else:
            print("Error getting data from topic {}".format(topic))
            mean_total_trial_t = None

        return mean_total_trial_t

    def has_data(self):
        if len(self._data) > 0:
            return True
        else:
            return False

    def parse_files(self, data):
        """ Rearranges the order of files stored in a dictionary associated with a specific training 
            day by specific ending index
            
        Parameters:
        -----------
        data        : (dict) Dictionary with dates for keys and file name paths for values
        
        Returns:
        -----------
        parsed_data : (dict) Dictionary with the same values associated with the keys but reararnged
        
        Example: 
        -----------
        TO-DO
                
        """
        parsed_data = {}
        for date_str, str_list in data.items():
            if date_str not in parsed_data:
                parsed_data[date_str] = None

            # Get times and file keys
            times_list = []
            keys_list = []
            for string in str_list:
                temp = re.findall('(\d+)(?=.)', string)
                times_list.append(temp[3])
                keys_list.append(temp[4])

            max_idx = [i for i, value in enumerate(keys_list) if value == max(keys_list)]
            time_match = times_list[max_idx[-1]]  # time substring files
            if self.verbose:
                print("Largest key found: {} from {}".format(keys_list[max_idx[-1]], time_match))

            # Fill new dictionary with filtered list of files
            parsed_data[date_str] = [j for j in str_list if time_match in j]

        print("Sorted files with most consecutive keys")
        return parsed_data

    def parse_ros2_topics(self, df):
        """ Reads a DataFrame and returns a Python dictionary with topics as keys and the 
        correspinding data type in the values associated with that key. This shrinks down the df 
        space since many Nan elements are created from non-overlapping ROS2 datatypes
        
        Parameters:
        ----------
        df          : (DataFrame) A pandas DataFrame containing the converted ROS2 bag file
        
        Return:
        --------- 
        data       : (dict) A Python dictionary where each key is a unique ROS2 topic and the value 
                     containing the associated data
        
        Example:
        ---------
        >> df
                               topic              time_ns  ... closed_file opened_file
        0                    /rosout  1700254464603345720  ...         NaN         NaN
        1       /robot/command/force  1700254464603372652  ...         NaN         NaN
        2          /parameter_events  1700254464603388062  ...         NaN         NaN
        3              /cursor/force  1700254464603401617  ...         NaN         NaN
        4                    /rosout  1700254464603573544  ...         NaN         NaN
        ...                      ...                  ...  ...         ...         ...
        417735         /cursor/force  1700255126708233535  ...         NaN         NaN
        417736  /robot/command/force  1700255126708379677  ...         NaN         NaN
        417737  /robot/command/force  1700255126708697736  ...         NaN         NaN
        417738  /robot/command/force  1700255126712198773  ...         NaN         NaN
        417739  /robot/command/force  1700255126712477497  ...         NaN         NaN

        [2639233 rows x 24 columns]
        
        >> agent = DataAgent()
        >> ros2_dict = agent.parse_ros2_topics(df)
        >> ros2_dict
        {'/robot/command/force':{'type':'/geometry_msgs/msg/Point', 
                                 'ts': [1700254464603372652, ...],
                                 'data': {'x':[0.0...], 'y':[0.0...], 'z':[0.0...]}
                                 'start_time': 2023-11-17 3:54:24 PM
                                 }, 
         '/rosout':{'type':'/rcl_interfaces/msg/Log', 
                    'ts': [1700254464603345720, ...],
                    'data': {'x':[0.0...], 'y':[0.0...], 'z':[0.0...]}
                    'start_time': '2023-11-17 3:54:24 PM'
                   },                 
        """
        print("Parsing ROS2 data from DataFrame")

        start_time = df['time_ns'].iloc[0]

        ros2_dict = {'info': '.mat file created using DataAgent', 'topics': [], 'samples': [],
                     'sample_rate': 'Check the samples field for sample rates of data channels',
                     'start_time': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(start_time / 1e9)),
                     'file_modified': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())}

        date = time.strftime('%Y_%m_%d', time.localtime(start_time / 1e9))
        ros2_dict['name'] = "{}_{}_rosbag_A_0".format(self.subject, date)

        for i, topic in enumerate(df['topic'].unique()):

            # Get all rows matching the topic name
            topic_df = df.loc[df.index[df['topic'].isin([topic])]]

            # Get start time datestamp , timestamp, and ROS2 topic data type
            temp = {'type': str(topic_df['type'].iloc[0]), 'topic': topic, 'time_index': topic_df['time_ns'].to_list(),
                    'time_s': ((topic_df['time_ns'] - start_time) / 1e9).to_list(), 'n_samples': len(topic_df)}
            # temp['start_time'] = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(start_time/1e9))
            temp['sample_rate'] = 1 / np.mean(np.diff(temp['time_s']))

            # Find the sub data types associated with the ros2 topic            
            subtype_keys = get_ros2_datatypes(topic_df['type'].iloc[0])

            # Fill in the 'data' 
            data = {}
            for key in subtype_keys:
                data[key] = topic_df[key].to_list()
            temp['msg_data'] = data

            # Need to remove the '/' character from the key names otherwise the .mat file wont save
            # topic = topic[1:] if topic[0] == '/' else topic
            ros2_dict['topics'].append(topic)
            ros2_dict['samples'].append(temp)

        return ros2_dict

    def read_bag(self, file_path, combine=True, display_topics=False):
        """ Opens and reads the bag file specified by the parameter input.

        Parameters:
        ----------
        file_path      : (str) Single string or list of strings to concatenate data from multiple 
                               files
        combine        : (bool) Optional argument to combine the data read from the bag files
        display_topics : (bool) Optional argument to display the topic data from the bag file(s)


        ----------
        Updates-> _data  : (list) list of Reader object(s) containing ROS2 bag file data from each 
                                file read
        
        Notes:
        ----------
        Some bag files can have broken indicies preventing them from being loaded. To fix them, 
        follow instructions from: https://stackoverflow.com/questions/18259692/how-to-recover-a-corrupt-sqlite3-database/57872238#57872238 
        
        """
        main_df = None
        self._data = []
        # Need to define the storage ID with the Reader object before attempting to access files 
        ft = self.file_type
        if ft == '.db3' or ft == 'db3' or ft == 'sqlite3':
            storage_id = 'sqlite3'
        elif ft == '.mcap' or ft == 'mcap':
            storage_id = 'mcap'

        if isinstance(file_path, str):
            files = [file_path]
        elif isinstance(file_path, list):
            files = file_path
        else:
            print('Error - Be sure to pass a string directory or list of string directories')
            return

        for f in file_path:
            print(" | Reading bag file from '{}'".format(f))
            temp = Reader(f, storage_id=storage_id)

            if display_topics:
                print("Topics present: ")
                print("{}\n".format([i for i in temp.topics]))

            # Convert the bag file into a DataFrame
            df_file = pd.DataFrame(temp.records)

            if combine:
                if main_df is not None:
                    main_df = pd.concat([main_df, df_file], ignore_index=True, axis=0)
                else:
                    main_df = df_file
            else:
                self._data.append(df_file)

        if combine:
            self._data = main_df

    def save_as_mat(self, file_path, df):
        """ Helper function to save the dataframe as a .mat file

        Parameters:
        ----------
        file_path : (str) File path to save the data to including the file name
        df        : (DataFrame) DataFrame object to save as a .mat file
        
        """
        if file_path[-4:] != '.mat':
            file_path = file_path + ".mat"

        print("Saving mat file to '{}'".format(file_path))
        ros_dict = self.parse_ros2_topics(df)
        sio.savemat(file_path, ros_dict)

    def save_as_htf5(self, df, file_name):
        """
        # http://docs.h5py.org/en/latest/high/file.html
        # http://blog.tremily.us/posts/HDF5/
        # http://www.sam.math.ethz.ch/~raoulb/teaching/PythonTutorial/data_storage.html

        """
        file_path = file_name + '.h5'
        print("saving dataframe to: '{}'".format(file_path))

        # Currently disorganized with the dataset saving. Need to investigate how to organize it better
        # df.to_hdf(file_name, key='/data', mode='w')

    def save_data(self, save_path=None, file_type='txt', df=None):
        """ Function that saves all the data collected from a bag/log file into a directory with the
        specified data format.
        
        Parameters:
        ------------
        save_path    : (str) Specified directory with the file name included
        file_type    : (str) File type to save the file as
        df           : (DataFrame) Optional input to save the specified DataFrame object isnead of the stored one
        """
        if True:
            # try:
            # print('\n=== Grabbing all data ===\n')
            # [folder_path, new_date] = self.build_path(date)
            # file_name = self.subject + '_' + new_date + '_DATA.' + file_type

            # Overwrite folder path if suggested
            if save_path is not None:
                folder_path = save_path

            if df is not None:
                data = df
            else:
                data = self._data

            if type(data) != list:
                data = [data]

            for dframe in data:
                if file_type == 'hdf5' or file_type == '.hdf5':
                    # print("Saving bag file data to:\n    {}".format( save_path + '.h5'))
                    self.save_as_htf5(dframe, save_path)
                elif file_type == 'h5' or file_type == '.h5':
                    self.save_as_htf5(dframe, save_path)
                elif file_type == 'csv' or file_type == '.csv':
                    save_as(dframe, save_path)
                elif file_type == 'txt' or file_type == '.txt':
                    save_as(dframe, save_path)
                elif file_type == 'mat' or file_type == '.mat':
                    self.save_as_mat(save_path, dframe)
        # except:
        #    print("Something went wrong with saving the data... Please investigate")

    def search_for_files(self, file_path=None, file_type=None, date_tag=None):
        """ Function that recursively goes through each folder from a received directory and returns a 
            python dictionary with the file names for each unique date detected. Days with multiple 
            files in them are numerically arranged.
        
        Parameters
        ----------
        file_path : (str) string containing an absolute directory to the data 
        file_type : (str) data file type
        date_tag  : (str) Argument to specify which dates to return in the data
            
        Return
        ---------
        data : (list) A List of Python dictionary with dates for keys and lists of associated filenames 
        		      for each unique date 
        		      
        """
        if file_path is not None and type(file_path) == str:
            if self.check_path(file_path):
                self.search_path = file_path

        if file_type is not None and type(file_type) == str:
            self.file_type = file_type

        print("Searching for bag files in '{}' ...".format(self.search_path))
        # 'items' contains at least 1 tuple with: ('search path', [list of folders], [list of files])
        items = [i for i in os.walk(self.search_path)]

        if len(items) == 1 and (items[0][1]) == 0 and (items[0][2]) == 0:
            print("No folders or files found in search path '{}' ".format(self.search_path))
            return None

        metadata = []

        # If any folders are present search them first as priority
        if len(items[0][1]) > 0:
            print("Found {} folder(s) in {}".format(len(items[0][1]), self.search_path))
            if self.verbose:
                for f in items[0][1]: print("|    {}".format(f))

            if date_tag: print("|    Filtering files using date tag '{}'".format(date_tag))

            for i, folder in enumerate(items[0][1]):
                data = {}

                # If Specific date was requested, only collect files with matching date
                if date_tag:
                    date_match, _ = parse_str_date_info(folder)
                    if date_match != date_tag:
                        continue

                file_search_path = str(self.search_path) + "/" + str(folder) + "/*" + str(self.file_type)
                if self.verbose: print("\nSearching for files in '{}'...".format(file_search_path))
                data['folder_path'] = items[i + 1][0]  # fill in the absolute path to the folder
                data['files'] = sorted(
                    glob.glob(file_search_path))  # Keep the full path for all files matching the file type, and sort

                if len(data['files']) > 0:
                    data['block'] = []
                    if self.verbose: print(
                        "|  Found {} files with file type '{}'".format(len(data['files']), self.file_type))
                    for f in data['files']:
                        if self.verbose: print("|    '{}'".format(f))
                        data['date_tag'], data['date'], data[
                            'file_tag'], block, date_info = parse_digits_from_string(f)
                        data['year'], data['month'], data['day'] = date_info
                        data['block'].append(
                            int(block))  # Add the block number to the list, easiest way to keep track of files
                        data['timestamp'] = convert_digits_to_timestamp(
                            data['file_tag'])  # Convert time into HH:MM:SS
                else:
                    print("No files ending in' {}' found in '{}'".format(self.file_type, data['folder_path']))
                    continue

                # Fill in metadata with remaining items and sort the block numbers
                data['file_type'] = self.file_type
                data['block'] = sorted(data['block'])

                if self.verbose:
                    print("\n folder metadata: ")
                    for m in data:
                        print("|   {}:".format(m))
                        if isinstance(data[m], list):
                            for j in data[m]:
                                print("|   |   {}".format(j))
                        else:
                            print("|   |   {}".format(data[m]))

                # Add folder data to data list
                metadata.append(data)

        # Local bag files in the search directory
        elif len(items[0][2]) > 0:
            data = {}
            file_search_path = str(self.search_path) + "/*" + str(self.file_type)
            data['files'] = glob.glob(file_search_path)
            if len(data['files']) > 0:
                data['block'] = []
                for f in data['files']:
                    if self.verbose: print("|    '{}'".format(f))
                    data['date_tag'], data['date'], data['file_tag'], block, date_info = parse_digits_from_string(
                        f)
                    data['year'], data['month'], data['day'] = date_info
                    data['block'].append(int(block))
                    data['timestamp'] = convert_digits_to_timestamp(data['file_tag'])
                    data['file_type'] = self.file_type
                    data['block'] = sorted(data['block'])

                    # Add file data to data list
                    metadata.append(data)

            else:
                print("No files ending in' {}' found in '{}'".format(self.file_type, self.search_path))

        # Sorting with the order of the files by the file_tag
        metadata = self.sort_by(metadata, 'file_tag')

        return metadata

    def set_file_type(self, file_type=None):
        """ Overwrite the file type to read from
        """
        if not file_type:
            print("Warning, no file type specified. Please enter the file type as '.db3' or '.mcap'")
        else:
            self.file_type = file_type

    def set_notebook_headers(self, args, output=False):
        self.notebook_headers = args.split(',')
        if self.verbose:
            print('Filling in Notebook options:')
            if self.notebook_headers is not None:
                for n in self.notebook_headers:
                    print("\t" + n)
        else:
            print("No notebook data columns specified...")
        if output:
            return self.notebook_headers

    def sort_by(self, file_list, key, descending=False):
        """ Helper function to rearrange the contents of a list in ascending (or descending) order
       with respect to a dictionary key
       
       Parameters:
       -----------
       file_list     : (list) A list of Python dictionaries which have the same key:value pairs
       key           : (str) The dictionary key to sort all other data by
       
       Return:
       -----------
       sorted_list   : (list) The list of dictionaries rearranged
       
       Example:
       ----------
       >> from data_utils import DataAgent
       >> data = [{'name': 'Yoda', 'age':900}, {'name': 'Grogu', 'age':50}]
       >> agent = DataAgent()
       >> sorted_list = agent.sort_by(data, 'age')
       >> sorted_list
       >> [{'name': 'Grogu', 'age': 50}, {'name': 'Yoda', 'age': 900}]
       """

        sorted_list = sorted(file_list, key=lambda x: x[key], reverse=descending)
        return sorted_list

    def save_file(self, df_file, folder_path, file_name, overwrite=False, default_file_type='txt'):
        """ Helper function that saves the file to the specific folder path

        Parameters:
        -----------
        df_file           : (DataFrame) Dataframe object to save
        folder_path       : (str) directory to save the file
        file_name         : (str) name of the file to save
        overwrite         : (bool) Option to overwrite the file if it already exists
        default_file_type : (str) Default file type to save as if not specified
        """

        # Check if file_name already has file type, if not use default
        matches = re.search('.\w+', file_name)
        if matches[-1][0] != '.':
            file_name = file_name + '.' + default_file_type

        if os.path.exists((folder_path + file_name)):
            if not overwrite:
                print('{} already in path. Set enable to True if you would like to save a new file'.format(
                    (folder_path + file_name)))
        else:
            print('Saving data to:\n ' + (folder_path + file_name))
            save_as(df_file, folder_path, file_name)

    def save_notebook(self, file_path, overwrite=False):
        """ Saves the loaded notebook to a specified file path

        Parameters:
        -----------
        file_name : (str) Name of the file to save
        file_path : (str) Directory to save the file
        overwrite : (bool) Option to overwrite the file if it already exists

        """

        if not os.path.exists(file_path):
            print("Error - Please enter a valid file path")
        else:
            if not overwrite:
                print('{} already in path. Set enable to True if you would like to save a new file'.format(file_path))
            else:
                print('Saving notebook to: "' + file_path + '"')
                self.notebook_file.save(file_path)

    def transfer_data(self, files=None, new_path=None, tag=None, move_parent_folder=False):
        """ Function that moves the files defined as filepath strings to a new directory.

        Parameters:
        -----------
        files              : (str/list) Filepath string or list of filepath strings to move, can be 
                               set to None to select files in the search directory
        new_path           : (str) replacement directory if not using the save path
        tag                : (str) String containing a date entry in the format of MM/DD/YY
        move_parent_folder : (bool) Option to move the folder containing the data
            
            
        Note:
        ------------
        Using shutil to transfer files or directories between disk drives has been known to thrown 
        errors. This can happen if write permissions aren't set for the new directory. Using 'rsync' 
        in a bash script would be a alternative method
        
        
        Example 1, move single file:
        ---------------------------
        >> search_path = '/home/user/documents'
        >> save_path = '/home/user/downloads'
        >> agent = DataAgent('TestUser', search_path, save_path)
        >>
        >> with open('/home/user/documents/myfile.txt', 'w') as fp: # create empty text file
        >> ...  pass
        >> agent.transfer_data('/home/user/documents/myfile.txt','/home/user/downloads/myfile.txt')

        >>
        >> import glob
        >> glob.glob(save_path)
           ['/home/user/downloads/TestUser/myfile.txt']
        
        
        Example 2, move and rename folder containing files in search path using tag
        ---------------------------
        >> date = '7/27/23' # Date tag entry in the format of MM/DD/YY
        >>
        >> import os
        >> os.mkdir('/home/user/documents/MyFolder')
        >> with open('/home/user/documents/MyFolder/myfile2.txt', 'w') as fp:
        >> ...  pass
        >> with open('/home/user/documents/MyFolder/myfile3.txt', 'w') as fp:
        >> ...  pass
        >>
        >> agent.transfer_data(None, save_path, date)
        >>
        >> glob.glob(search_path)
           ['/home/user/documents']
        >> glob.glob(save_path)
           ['/home/user/downloads/TestUser/TestUser_2023_07_27'/myfile2.txt', 
            '/home/user/downloads/TestUser/TestUser_2023_07_27'/myfile3.txt']
                    
        """

        if not (new_path or tag):
            print("Error - Please specify either new directory or tag")
            return
        elif not new_path:
            print("Error - Please specify at least a new directory")
            return

        if not os.path.isdir(new_path):
            print("Error - Please enter valid new path")
            return

        if files:

            file_list = (files if type(files) is list else [files])
        else:
            # Get all the files located in the search path if no file is provided
            file_list = glob.glob(self.search_path + "/*")
            if not len(file_list) > 0:
                print('Error - No files or folders found in search directory')
                return

        # If true, grabbing the parent directory of the file
        if move_parent_folder:
            print("'move_parent_folder' set to True, moving all contents in parent folder")
            file_list = [str(os.path.dirname(file_list[0]))]

        if self.verbose:
            print("Preparing data to transfer: ")
            [print(" | [{}]".format(i)) for i in file_list]

        folder_path = os.path.join(new_path, self.subject)
        if not os.path.isdir(folder_path):
            os.makedirs(folder_path)

        if tag:
            [_, new_date] = self.build_path(tag)
            folder_path = os.path.join(folder_path, self.subject + '_' + new_date)

        for file_ in file_list:
            if os.path.isfile(file_):
                file_name = os.path.basename(file_)
                transfer_path = os.path.join(folder_path, file_name)
                if os.path.isfile(transfer_path):
                    print(" | | Stopping transfer of {} to {}, file already exists".format(file_, transfer_path))
                    continue
            if os.path.isdir(file_):
                transfer_path = folder_path
                if os.path.isdir(transfer_path):
                    print(" | | Stopping transfer of {} to {}, folder already exists".format(file_, transfer_path))
                    continue

            if self.verbose:
                print("Moving file:\n | Original path: {}\n | New path {}".format(file_, transfer_path))

            # If not sudo, using the 'shutil.move()' method throws two error messages: [Errno 18, Errno 95]
            try:
                if os.path.isdir(transfer_path):
                    print("Deleting directory in 60 seconds. Make sure the folder has been successfully transferred")
                    time.sleep(60)
                    shutil.move(os.path.realpath(file_), transfer_path)
            except Exception as e:
                print(e)
                print("Error - Could not move file to new directory")

    def write_to_sheet(self, sheet_name, rowx, colx, value):
        """ Helper function to write to a specific cell in the sheet

        Parameters:
        -----------
        sheet_name : (str) Name of the sheet to write to
        rowx       : (int) Row index to write to
        colx       : (int) Column index to write to
        value      : (str) Value to write to the cell
        """
        if self.verbose:
            print(" - Writing {} to sheet '{}' at row {}, column {}".format(value, sheet_name, rowx, colx))
        if isinstance(self.notebook_file, xlrd.book.Book):
            ref_sheet = self.notebook_file.sheet_by_name(sheet_name)
            xf_index = ref_sheet.cell_xf_index(rowx, colx)
            # wtsheet = self.get_sheet_by_name(self.notebook_copy, sheet_name)
            wtsheet = self.get_sheet_by_name(self.notebook_file, sheet_name)
            wtsheet.write(rowx, colx, value, self.notebook_style_list[xf_index])

        elif isinstance(self.notebook_file, openpyxl.workbook.workbook.Workbook):
            wksheet = self.notebook_file[sheet_name]
            # wksheet = self.notebook_copy[sheet_name]
            try:
                wksheet.cell(row=rowx, column=colx, value=value)
            except Exception as e:
                print(" - Error writing to sheet: {}. Please make sure that header name exists".format(e))

    def get_sheet_by_name(self, book, name):
        """Helper function to get a sheet by name from xlwt.Workbook, a strangely missing method.

        Parameters:
        -----------
        book : (xlwt.Workbook) The workbook to search
        name : (str) The name of the sheet to get

        Return:
        --------
        sheet or None if no sheet with the given name is present.
        """
        # Note, we have to use exceptions for flow control because the
        # xlwt API is broken and gives us no other choice.
        try:
            for sheet in book._Workbook__worksheets:
                if sheet.name == name:
                    return sheet
        except IndexError:
            return None
