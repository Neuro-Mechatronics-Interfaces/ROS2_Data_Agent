"""
 Copyright 2024 Carnegie Mellon University Neuromechatronics Lab

 This Source Code Form is subject to the terms of the Mozilla Public
 License, v. 2.0. If a copy of the MPL was not distributed with this
 file, You can obtain one at https://mozilla.org/MPL/2.0/.

 Contact: Jonathan Shulgach (jshulgac@andrew.cmu.edu)

 This script updates a workbook/spreadsheet with experiment data without changing formatting.
"""
import os
import xlwt
import openpyxl
from data_agent import DataAgent


def create_demo_file_xlsx():
    # Create a xlsx file using openpyxl, and rename the sheet to 'First'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'First'
    #colors = 'white black red green blue pink turquoise yellow'.split()
    # Creating a list of the colors but using aRGB hex values
    colors = ['00FFFFFF', '00000000', '00FF0000', '0000FF00', '000000FF', '00FF00FF', '0000FFFF', '00FFFF00']
    fancy_styles = [openpyxl.styles.NamedStyle(
        name=colour, font=openpyxl.styles.Font(bold=True, italic=True, color=colour)) for colour in colors]
    for rowx in range(len(colors)):
        ws.cell(row=rowx+1, column=1, value=rowx)
        ws.cell(row=rowx+1, column=2, value=colors[rowx]).style = fancy_styles[rowx]
    wb.save('demo_in.xlsx')


def create_demo_file_xls():
    # Create an input file for the demo
    wtbook = xlwt.Workbook()
    wtsheet = wtbook.add_sheet(u'First')
    colors = 'white black red green blue pink turquoise yellow'.split()
    fancy_styles = [xlwt.easyxf(
        'font: name Times New Roman, italic on;'
        'pattern: pattern solid, fore_colour %s;'
        % colour) for colour in colors]
    for rowx in range(len(colors)):
        wtsheet.write(rowx, 0, rowx)
        wtsheet.write(rowx, 1, colors[rowx], fancy_styles[rowx])
    wtbook.save('demo_in.xls')


if __name__ == '__main__':

    # First lets create a demo workbook with some data. We will edit 8 cells where each one has a distinct color and
    # italicized text, then save it to the local directory. Feel free to open the workbook to make sure it is filled
    # (and was created correctly!)
    # create_demo_file_xls # Keeping this here as a reference for creating a .xls file with the same data
    create_demo_file_xlsx()

    # Now let's create a DataAgent object to handle the data, passing in a save path to the current directory
    agent = DataAgent(save_path=os.getcwd())

    # Load the workbook. We also want to be able to keep the original formatting styles the workbook has
    agent.load_notebook('demo_in.xlsx')

    # Now let's change data content on the sheet 'First' without changing the formatting
    # ('pink' -> 'MAGENTA', 'turquoise' -> 'CYAN')
    fixups = [(5, 1, 'MAGENTA'), (6, 1, 'CYAN'),(7,1,'YELLOW')]
    for rowx, colx, value in fixups:
        agent.write_to_sheet('First', rowx, colx, value)

    # Lastly, save the workbook as a new file
    agent.save_notebook('demo_out.xlsx')
